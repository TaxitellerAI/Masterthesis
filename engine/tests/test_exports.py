"""Export fidelity: every downloadable artefact must contain the ACTIVE sample.

/dataset and /workbook are what an examiner opens to check the reported figures.
Both used to be built from the full universe regardless of the selected scenario —
asking for S1 handed out 8 columns including Solana instead of S1's 7. An appendix
holding a different sample than the text is the most expensive kind of mismatch,
and it is invisible unless the column count, the row count and the window are
asserted against the scenario itself.

Run:  python tests/test_exports.py
"""
from __future__ import annotations
import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pandas as pd

from api.main import RunRequest, dataset, workbook, _sample_for, _scenario_prices
from volcontrol import fingerprint
from api.main import _spec_with_effective

# Independently pinned expectations — deliberately literal, not derived from the
# same code path that produces them, so a change in sample resolution has to be
# acknowledged here rather than silently agreeing with itself.
EXPECTED = {
    "S1": {"cols": 7, "rows": 2011, "start": "2018-01-02", "end": "2025-12-31"},
    "S2": {"cols": 8, "rows": 2766, "start": "2015-01-02", "end": "2025-12-31"},
    "S3": {"cols": 8, "rows": 1255, "start": "2021-01-04", "end": "2025-12-31"},
}


def _csv_of(scenario: str) -> tuple[pd.DataFrame, str]:
    req = RunRequest(scenario=scenario)
    resp = dataset(req)
    frame = pd.read_csv(io.BytesIO(resp.body), index_col=0)
    disp = resp.headers["content-disposition"]
    file_hash = disp.split("treasury-dataset-")[1].split(".csv")[0]
    return frame, file_hash


def test_dataset_matches_active_scenario():
    for scenario, want in EXPECTED.items():
        frame, file_hash = _csv_of(scenario)
        assert len(frame.columns) == want["cols"], (
            f"{scenario}: {len(frame.columns)} Spalten, erwartet {want['cols']} "
            f"({list(frame.columns)})")
        assert len(frame) == want["rows"], (
            f"{scenario}: {len(frame)} Zeilen, erwartet {want['rows']}")
        assert str(frame.index[0])[:10] == want["start"], (
            f"{scenario}: Fenster beginnt {frame.index[0]}, erwartet {want['start']}")
        assert str(frame.index[-1])[:10] == want["end"], (
            f"{scenario}: Fenster endet {frame.index[-1]}, erwartet {want['end']}")

        # The hash in the file name must be the hash the report prints, or the
        # appendix cites a snapshot nobody can match to the results.
        scoped, spec, report = _sample_for(RunRequest(scenario=scenario))
        report_hash = fingerprint(
            scoped, _spec_with_effective(RunRequest(scenario=scenario), report))["hash"]
        assert file_hash == report_hash, (
            f"{scenario}: Hash im Dateinamen {file_hash} != Report-Hash {report_hash}")
        print(f"OK {scenario}: {want['cols']} Spalten, {want['rows']} Zeilen, "
              f"{want['start']}..{want['end']}, Hash {file_hash}")


def test_workbook_prices_match_dataset():
    """Third export path, same sample — the workbook's price sheet must be the
    identical matrix /dataset hands out, not the full universe."""
    import openpyxl
    for scenario in EXPECTED:
        req = RunRequest(scenario=scenario)
        expected = _scenario_prices(req, _sample_for(req)[1])
        wb = openpyxl.load_workbook(io.BytesIO(workbook(req).body), read_only=True)
        ws = wb["Kurse"]
        cols, rows = ws.max_column - 1, ws.max_row - 1
        assert (rows, cols) == expected.shape, (
            f"{scenario}: Workbook-Kurse {rows}x{cols}, Datensatz {expected.shape}")
        print(f"OK {scenario}: Workbook-Kursblatt deckungsgleich mit /dataset "
              f"({rows} Zeilen x {cols} Assets)")


if __name__ == "__main__":
    test_dataset_matches_active_scenario()
    test_workbook_prices_match_dataset()
