"""Transparency guarantee test: the Excel formulas in the exported workbook must
reproduce the engine's metrics EXACTLY. Builds the workbook, recalculates it with
headless LibreOffice, and compares the Kennzahlen sheet against the engine.

Runs on the REPORTED configuration — frozen snapshot, sample design S1, chained
risk-free series — not on a synthetic fixture. The earlier synthetic version proved
only that the formulas were self-consistent on data nobody reports (its Buy-and-Hold
drawdown was 0.1863 against 0.1592 on S1), so a formula error that only shows up on
the real sample, the rf series, or the point-in-time sleeve would have passed.

Skips (with a clear message) if LibreOffice is not installed — the guarantee is
then only checked on machines that can recalculate formulas.

Run:  python tests/test_workbook.py
"""
from __future__ import annotations
import os
import shutil
import subprocess
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from volcontrol import run_strategies, portfolio_weights, build_workbook, fingerprint
from volcontrol.backtest import _blended_cost_bps

# The test drives the API's OWN preparation path (sample resolution, chained rf,
# monthly weight rebalancing, point-in-time sleeve). Rebuilding that config by hand
# here would let the test and the endpoint drift apart silently — which is exactly
# how the previous synthetic version stopped testing anything that gets reported.
from api.main import RunRequest, _prepared, _scenario_prices

SOFFICE = shutil.which("soffice") or (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists("/Applications/LibreOffice.app/Contents/MacOS/soffice") else None
)

TOL = 5e-4  # metrics agree to 4 decimals (rounding of exposure values in the sheet)


def test_workbook_reproduces_engine():
    if not SOFFICE:
        print("SKIP — LibreOffice (soffice) nicht gefunden; Formel-Recalc nicht prüfbar.")
        return

    # Exactly what the thesis reports: frozen EUR snapshot, S1, chained rf.
    req = RunRequest(scenario="S1")
    crypto_share, target_vol = req.crypto_share, req.target_vol
    rets, cfg, rf_info, spec, sample_report, pit = _prepared(req)
    prices = _scenario_prices(req, spec)
    assert len(rets) == 2010, f"S1 sollte 2010 Renditezeilen haben, hat {len(rets)}"
    assert cfg.weight_rebalance == "monthly", cfg.weight_rebalance
    assert cfg.rf_series is not None, "rf-Reihe fehlt — Test liefe auf konstantem rf"

    run = run_strategies(rets, cfg, crypto_share, pit_builder=pit)
    key = f"VolControl_{int(target_vol * 100)}"
    engine_bh = run["strategies"]["BuyHold"]
    engine_vc = run["strategies"][key]

    meta = {
        "crypto_share": crypto_share, "target_vol": target_vol,
        "base_currency": "EUR", "source": "frozen", "scenario": "S1",
        "cost_bps": _blended_cost_bps(crypto_share, cfg),
        "fingerprint": fingerprint(rets, {"scenario": "S1"}),
        "trad_split": {"MSCI_World": 0.6, "Global_Bonds": 0.3, "Gold": 0.1},
        "trad_is_base": True, "rf_mode": rf_info.get("mode", "estr_chained")
        if isinstance(rf_info, dict) else "estr_chained", "generated_at": "test",
    }
    weights = portfolio_weights(crypto_share, list(rets.columns), cfg)
    xbytes = build_workbook(prices, rets, weights, engine_vc["exposure"],
                            engine_vc["returns"], cfg, meta, {})

    with tempfile.TemporaryDirectory() as td:
        src = os.path.join(td, "wb.xlsx")
        with open(src, "wb") as f:
            f.write(xbytes)
        subprocess.run(
            [SOFFICE, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", td, src],
            check=True, capture_output=True, timeout=180,
        )
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(td, "wb.xlsx"), data_only=True)
        ws = wb["Kennzahlen"]
        # row 2 = Buy-and-Hold, row 3 = selected vol-control; columns B..G
        excel = {
            "bh": [ws.cell(2, c).value for c in range(2, 8)],
            "vc": [ws.cell(3, c).value for c in range(2, 8)],
        }

    def _check(tag, cells, eng):
        expected = [eng["ann_return"], eng["cagr"], eng["ann_vol"],
                    eng["sharpe"], eng["max_drawdown"], eng["cvar_95"]]
        names = ["ann_return", "cagr", "ann_vol", "sharpe", "max_drawdown", "cvar_95"]
        for name, got, want in zip(names, cells, expected):
            assert isinstance(got, (int, float)), f"{tag}.{name}: Formel ergab {got!r}"
            assert abs(got - want) < TOL, f"{tag}.{name}: Excel {got:.6f} vs Engine {want:.6f}"

    _check("BuyHold", excel["bh"], engine_bh)
    _check(key, excel["vc"], engine_vc)
    print(f"OK — Excel-Formeln reproduzieren die Engine-Kennzahlen auf S1 "
          f"(n={len(rets)}, Toleranz {TOL}); BH & {key} über 6 Metriken geprüft. "
          f"BuyHold MaxDD {engine_bh['max_drawdown']:.4f}, "
          f"Sharpe {engine_bh['sharpe']:.4f}.")


if __name__ == "__main__":
    test_workbook_reproduces_engine()
