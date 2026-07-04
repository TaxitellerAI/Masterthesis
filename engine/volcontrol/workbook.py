"""Transparency workbook: an .xlsx where the raw prices are data and every
downstream value is a LIVE Excel formula. An examiner can change a price and watch
the returns, the portfolio, the wealth path and all closed-form metrics recompute —
the calculation is auditable end to end, not just reported.

Design:
  * Kurse      — raw (currency-converted) prices, values.
  * Renditen   — daily simple returns as formulas =Kurse!P_t/P_{t-1}-1.
  * Gewichte   — the renormalised portfolio weights.
  * Portfolio  — Buy-and-Hold return =SUMPRODUCT(returns, weights); wealth, peak,
                 drawdown as formulas; the selected vol-control strategy composed
                 from the engine's exposure via =expo*r + (1-expo)*rf - cost.
  * Kennzahlen — ann. return, CAGR, vol, Sharpe, MaxDD, CVaR as formulas over the
                 Portfolio ranges (the crown jewel: verifiable metrics).
  * Deskriptiv / Sweep / Hypothesen / Regime / Walk-Forward — engine values, with a
    note that the stochastic tests (bootstrap/HAC) are computed in Python.
  * Info       — parameters, data fingerprint, and the exact formula for each metric.

The vol-control DAILY return depends on the engine's exposure logic, so exposure is
provided as values; everything computed FROM it is a transparent formula.
"""
from __future__ import annotations
from io import BytesIO
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .config import EngineConfig

_HEAD = Font(bold=True, color="0A1D3A")
_MUTE = Font(color="55606F")
_TITLE = Font(bold=True, size=14, color="0A1D3A")
_FILL = PatternFill("solid", fgColor="EAF1FA")
_HAIR = Border(bottom=Side(style="thin", color="C6CDD7"))
_PCT = "0.00%"
_NUM4 = "0.0000"


def _col(i: int) -> str:
    return get_column_letter(i)


def build_workbook(prices: pd.DataFrame, returns: pd.DataFrame, weights: dict,
                   exposure: pd.Series, vc_returns: pd.Series, cfg: EngineConfig,
                   meta: dict, extras: dict) -> bytes:
    wb = Workbook()

    assets = list(returns.columns)
    na = len(assets)
    rf_daily = cfg.rf_daily
    td = cfg.trading_days
    cost_bps = meta["cost_bps"]

    # Grid = the price dates the returns were built from (align every sheet to it).
    grid = list(returns.index)          # returns index (prices minus first row)
    price_grid = list(prices.index)
    n = len(price_grid)

    # weight vector renormalised to 1 over the present columns (as buy_and_hold does)
    wv = np.array([weights.get(c, 0.0) for c in assets], float)
    wv = wv / wv.sum() if wv.sum() else wv

    # position of the first vol-control observation on the price grid
    vc_start_date = vc_returns.index[0]
    exp_map = {d: float(v) for d, v in exposure.items()}

    # ── Info ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Info"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 62
    ws["A1"] = "Volatility-Control Treasury — Transparenz-Arbeitsmappe"
    ws["A1"].font = _TITLE
    rows = [
        ("", ""),
        ("Parameter", ""),
        ("Krypto-Quote", f"{meta['crypto_share']*100:.1f} %"),
        ("Zielvolatilität", f"{meta['target_vol']*100:.0f} %"),
        ("Traditionelle Aufteilung", ", ".join(f"{k} {v*100:.0f}%" for k, v in meta.get("trad_split", {}).items())
            + ("  (Thesis-Basisfall)" if meta.get("trad_is_base", True) else "  (≠ Thesis-Basisfall)")),
        ("Basiswährung", meta["base_currency"]),
        ("Risikofreier Zins p.a. (effektiv)", f"{cfg.rf_annual*100:.2f} %"
            + ("  (€STR-Fensterdurchschnitt, ECB)" if meta.get("rf_mode") == "estr" else "  (manuell konstant)")),
        ("Volatilitäts-Schätzer", cfg.vol_method),
        ("Rebalancing", cfg.rebalance),
        ("Exposure-Totband", f"{cfg.dead_band:.2f}" if cfg.dead_band else "aus"),
        ("Handelstage/Jahr (Annualisierung)", td),
        ("Transaktionskosten (bps, gemischt)", round(cost_bps, 2)),
        ("Datenquelle", meta["source"]),
        ("Daten-Hash (Fingerprint)", meta["fingerprint"]["hash"]),
        ("Datenfenster", f"{meta['fingerprint']['start']} – {meta['fingerprint']['end']}"),
        ("Beobachtungen (Renditen)", len(grid)),
        ("Assets", ", ".join(assets)),
        ("", ""),
        ("Formeln & Methodik (live nachrechenbar)", ""),
        ("Tagesrendite", "r_t = Kurs_t / Kurs_{t-1} − 1  (Blatt 'Renditen')"),
        ("Portfolio (Buy-and-Hold)", "r_p = Σ_i w_i · r_i  (SUMPRODUCT, Blatt 'Portfolio')"),
        ("Vol-Control-Rendite", "r_vc = e·r_p + (1−e)·rf − Kosten   (e = Exposure der Engine)"),
        ("Kosten", "Kosten_t = |e_t − e_{t-1}| · bps/10000"),
        ("Ann. Rendite (arithm.)", "Ø(r) · 252"),
        ("CAGR (geometrisch)", "(Π(1+r))^(252/N) − 1"),
        ("Ann. Volatilität", "StdAbw(r) · √252"),
        ("Sharpe Ratio", "(Ø(r) − rf) / StdAbw(r) · √252"),
        ("Max Drawdown", "min( Wealth_t / laufendes Maximum − 1 )"),
        ("CVaR 95 %", "Ø( r | r ≤ 5%-Quantil )  (Expected Shortfall)"),
        ("", ""),
        ("Hinweis", "Bootstrap-, Wilcoxon- und HAC-Tests sind stochastisch und werden in "
                    "Python berechnet (Blatt 'Hypothesen'); sie sind mit fixem Seed reproduzierbar, "
                    "aber nicht als Excel-Formel darstellbar."),
        ("Erzeugt", meta["generated_at"]),
        ("", "Kein Anlageratschlag — akademisches Nachvollziehbarkeits-Dokument."),
    ]
    r = 2
    for a, b in rows:
        ws.cell(r, 1, a).font = _HEAD if b == "" and a and not a[0].isdigit() else _MUTE
        ws.cell(r, 2, b)
        r += 1

    # ── Kurse (prices, values) ───────────────────────────────────────────────
    wsk = wb.create_sheet("Kurse")
    wsk.cell(1, 1, "Datum").font = _HEAD
    for j, a in enumerate(assets):
        wsk.cell(1, j + 2, a).font = _HEAD
    for i, d in enumerate(price_grid):
        wsk.cell(i + 2, 1, pd.Timestamp(d).strftime("%Y-%m-%d"))
        for j, a in enumerate(assets):
            v = prices.iloc[i][a]
            wsk.cell(i + 2, j + 2, None if pd.isna(v) else float(v))
    wsk.freeze_panes = "B2"

    # ── Renditen (formulas from Kurse) ───────────────────────────────────────
    wsr = wb.create_sheet("Renditen")
    wsr.cell(1, 1, "Datum").font = _HEAD
    for j, a in enumerate(assets):
        wsr.cell(1, j + 2, a).font = _HEAD
    for i, d in enumerate(price_grid):
        wsr.cell(i + 2, 1, pd.Timestamp(d).strftime("%Y-%m-%d"))
        if i == 0:
            continue
        for j in range(na):
            c = _col(j + 2)
            rr, rp = i + 2, i + 1
            cell = wsr.cell(i + 2, j + 2)
            cell.value = f'=IF(OR(Kurse!{c}{rr}="",Kurse!{c}{rp}=""),"",Kurse!{c}{rr}/Kurse!{c}{rp}-1)'
            cell.number_format = _PCT
    wsr.freeze_panes = "B2"

    # ── Gewichte (weights; horizontal band for SUMPRODUCT) ───────────────────
    wsg = wb.create_sheet("Gewichte")
    wsg.cell(1, 1, "Asset").font = _HEAD
    wsg.cell(2, 1, "Gewicht").font = _HEAD
    for j, a in enumerate(assets):
        wsg.cell(1, j + 2, a)
        wsg.cell(2, j + 2, float(wv[j])).number_format = _PCT
    wsg.cell(3, 1, "Summe").font = _MUTE
    last = _col(na + 1)
    wsg.cell(3, 2, f"=SUM(B2:{last}2)").number_format = _PCT

    # ── Portfolio (formulas) ─────────────────────────────────────────────────
    wsp = wb.create_sheet("Portfolio")
    heads = ["Datum", "BH-Rendite", "BH-Wealth", "BH-Peak", "BH-Drawdown",
             "VC-Exposure", "VC-Kosten", "VC-Rendite", "VC-Wealth", "VC-Peak", "VC-Drawdown"]
    for j, h in enumerate(heads):
        wsp.cell(1, j + 1, h).font = _HEAD
    rlast = _col(na + 1)  # last asset column letter in Renditen
    g_first, g_last = 2, na + 1
    grid_pos = {d: k for k, d in enumerate(grid)}
    vc_excel_row = None
    for k, d in enumerate(grid):
        row = k + 3  # returns start at price-grid row 3 (2nd date); align Portfolio same
        wsp.cell(row, 1, pd.Timestamp(d).strftime("%Y-%m-%d"))
        rr = row  # Renditen row for this date (same grid, row offset +2 from index -> here matches)
        # BH return = Σ w_i·r_i, with empty (non-traded) cells treated as 0 — this
        # mirrors the engine's skipna aggregation and never yields #VALUE!.
        rng = f"Renditen!{_col(g_first)}{rr}:{_col(g_last)}{rr}"
        b = wsp.cell(row, 2)
        b.value = f"=SUMPRODUCT(IF(ISNUMBER({rng}),{rng},0),Gewichte!$B$2:${last}$2)"
        b.number_format = _PCT
        # Wealth / Peak / Drawdown
        w = wsp.cell(row, 3)
        w.value = f"=(1+B{row})" if k == 0 else f"=C{row-1}*(1+B{row})"
        w.number_format = _NUM4
        p = wsp.cell(row, 4)
        p.value = f"=C{row}" if k == 0 else f"=MAX(D{row-1},C{row})"
        p.number_format = _NUM4
        dd = wsp.cell(row, 5)
        dd.value = f"=C{row}/D{row}-1"
        dd.number_format = _PCT
        # VC exposure (value from engine)
        e = wsp.cell(row, 6)
        e.value = round(exp_map.get(d, 0.0), 6)
        e.number_format = _NUM4
        # VC cost = |Δexposure| * bps
        co = wsp.cell(row, 7)
        co.value = (f"=ABS(F{row})*{cost_bps}/10000" if k == 0
                    else f"=ABS(F{row}-F{row-1})*{cost_bps}/10000")
        co.number_format = _NUM4
        # VC return = e*BH + (1-e)*rf - cost
        vr = wsp.cell(row, 8)
        vr.value = f"=F{row}*B{row}+(1-F{row})*{rf_daily}-G{row}"
        vr.number_format = _PCT
        if d == vc_start_date:
            vc_excel_row = row
        if vc_excel_row is not None and row >= vc_excel_row:
            vw = wsp.cell(row, 9)
            vw.value = f"=(1+H{row})" if row == vc_excel_row else f"=I{row-1}*(1+H{row})"
            vw.number_format = _NUM4
            vp = wsp.cell(row, 10)
            vp.value = f"=I{row}" if row == vc_excel_row else f"=MAX(J{row-1},I{row})"
            vp.number_format = _NUM4
            vdd = wsp.cell(row, 11)
            vdd.value = f"=I{row}/J{row}-1"
            vdd.number_format = _PCT
    wsp.freeze_panes = "B2"

    bh_first, bh_last = 3, len(grid) + 2
    vc_first = vc_excel_row or bh_first
    bh_ret = f"Portfolio!$B${bh_first}:$B${bh_last}"
    vc_ret = f"Portfolio!$H${vc_first}:$H${bh_last}"
    bh_dd = f"Portfolio!$E${bh_first}:$E${bh_last}"
    vc_dd = f"Portfolio!$K${vc_first}:$K${bh_last}"
    bh_wealth_last = f"Portfolio!$C${bh_last}"      # = Π(1+r), wealth base 1
    vc_wealth_last = f"Portfolio!$I${bh_last}"

    # ── Kennzahlen (formulas) ────────────────────────────────────────────────
    wsm = wb.create_sheet("Kennzahlen")
    mheads = ["Strategie", "Ann. Rendite", "CAGR", "Ann. Vol", "Sharpe", "Max Drawdown", "CVaR 95 %"]
    for j, h in enumerate(mheads):
        wsm.cell(1, j + 1, h).font = _HEAD

    def _metric_row(row, label, ret, dd, wealth_last):
        wsm.cell(row, 1, label)
        wsm.cell(row, 2, f"=AVERAGE({ret})*{td}").number_format = _PCT
        # CAGR from the compounded final wealth (avoids PRODUCT-array pitfalls).
        wsm.cell(row, 3, f"={wealth_last}^({td}/COUNT({ret}))-1").number_format = _PCT
        wsm.cell(row, 4, f"=STDEV({ret})*SQRT({td})").number_format = _PCT
        wsm.cell(row, 5, f"=(AVERAGE({ret})-{rf_daily})/STDEV({ret})*SQRT({td})").number_format = "0.000"
        wsm.cell(row, 6, f"=MIN({dd})").number_format = _PCT
        # Legacy PERCENTILE (no .INC) so both Excel and LibreOffice resolve it.
        wsm.cell(row, 7, f'=AVERAGEIF({ret},"<="&PERCENTILE({ret},{cfg.cvar_alpha}))').number_format = _PCT

    _metric_row(2, "Buy-and-Hold", bh_ret, bh_dd, bh_wealth_last)
    _metric_row(3, f"Vol-Control {int(meta['target_vol']*100)} %", vc_ret, vc_dd, vc_wealth_last)
    wsm.cell(5, 1, "Alle Kennzahlen sind live aus dem Blatt 'Portfolio' berechnet.").font = _MUTE

    # ── Engine-value sheets (Deskriptiv / Sweep / Hypothesen / Regime / WF) ──
    def _table(name, records, cols=None):
        w = wb.create_sheet(name)
        if not records:
            w.cell(1, 1, "keine Daten")
            return
        cols = cols or list(records[0].keys())
        for j, c in enumerate(cols):
            w.cell(1, j + 1, c).font = _HEAD
        for i, rec in enumerate(records):
            for j, c in enumerate(cols):
                v = rec.get(c)
                w.cell(i + 2, j + 1, v if not isinstance(v, float) else round(v, 6))
        w.freeze_panes = "A2"

    _table("Deskriptiv", extras.get("describe", []))
    _table("Sweep", extras.get("sweep", []))
    _table("Regime", extras.get("subperiods", []))
    _table("Walk-Forward", extras.get("walk_forward_folds", []))

    # Hypotheses is a nested dict -> flatten to key/value rows.
    wh = wb.create_sheet("Hypothesen")
    wh.cell(1, 1, "Kennzahl").font = _HEAD
    wh.cell(1, 2, "Wert").font = _HEAD
    rr = 2
    for k, v in extras.get("hypotheses", {}).items():
        wh.cell(rr, 1, k).font = _HEAD
        if isinstance(v, dict):
            rr += 1
            for kk, vv in v.items():
                wh.cell(rr, 1, f"   {kk}")
                wh.cell(rr, 2, round(vv, 6) if isinstance(vv, float) else vv)
                rr += 1
        else:
            wh.cell(rr, 2, round(v, 6) if isinstance(v, float) else v)
            rr += 1
    wh.cell(rr + 1, 1, "Bootstrap/Wilcoxon/HAC in Python berechnet (fixer Seed, reproduzierbar).").font = _MUTE

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
